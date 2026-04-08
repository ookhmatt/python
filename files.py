import csv
from argparse import ArgumentError

from xml.etree.ElementTree import Element, ElementTree, SubElement, parse

from openpyxl.workbook import Workbook

from lib.person import data_ur_from_pesel, bmi
from lib.validators import validate_pesel, isLiczba


def pokaz_auta():
    p = parse(r'files/auta.xml') # otwarcie istniejącego pliku, jak ni ema to będzie błąd
    r = p.getroot() # znalezienie głownego znacznika (rodzica), jak błąd to oznacza, że plik jest
    # pusty lub ma błedną strukturę
    print('Pojazdy:')
    for i, a in enumerate(r.findall('auto'), start=1):
        marka, model, rej = a.find('marka').text, a.find('model').text, a.get('rej')
        print(f"{i} -> {marka}, {model} -> {rej}")
    return r


def demo_files():
    while True:
        print('1 - zapis txt, csv')
        print('2 - odczyt txt, csv')
        print('3 - zapis xml')
        print('4 - odczyt xml')
        print('5 - aktualizacja xml')
        print('6 - usuwanie xml')
        print('7 - pickle')
        print('8 - shelve')
        print('9 - excel')
        print('k - koniec')
        odp = input('Co robimy? -> ')
        match odp:
            case '1':
                warn = 'nie umiesz czytać, debilu?'
                imie = input('Podaj imie -> ')
                if len(imie) > 0:
                    nazwisko = input('Podaj nazwisko -> ')
                    if len(nazwisko) > 0:
                        pesel = input('Podaj pesel -> ')
                        if validate_pesel(pesel):
                            waga = input('Podaj wagę w kg -> ')
                            if waga.isdigit():
                                wzrost = input('podaj wzrost w metrach -> ')
                                wzrost = wzrost.replace(',', '.')
                                if isLiczba(wzrost):
                                    with open(r"files\dane_osobowe.csv", 'a', encoding='utf-8') as f:
                                        print(imie, nazwisko, pesel, waga, wzrost, sep=';', file=f)

                                    with open(r'files\dane_osobowe_csv.csv', 'a', encoding='utf-8') as f:
                                        csv_writer = csv.writer(f)
                                        csv_writer.writerow([imie, nazwisko, pesel, waga, wzrost])
                                else: print('Antoni świrze, gdzie sa Caracale? :)')
                            else: print('serio?')
                        else: print('podaj dobry pesel, wariacie')
                    else: print(warn)
                else: print(warn)
            case '2':
                with open(r'files\dane_osobowe.csv', 'r', encoding='utf-8') as f:
                    # f.readline() # odczyt nagłowka pliku
                    for o in f:
                        dane = o.split(';')
                        print(dane)
                        print(f'{dane[0]} {dane[1]} ur: {data_ur_from_pesel(dane[2])} '
                              f'ma bmi: {bmi(int(dane[3]), float(dane[4].strip('\n'))):.2f}')

                # z biblioteką csv
                with open(r'files\dane_osobowe.csv', 'r', encoding='utf-8') as f:
                    csv_reader = csv.reader(f)
                    # nagłówki = next(csv_reader) # jeśli sa to je w ten sposób odczytujemy
                    for rekord in csv_reader:
                        rekord = rekord[0].split(';')
                        print(f'{rekord[0]} {rekord[1]} ur: {data_ur_from_pesel(rekord[2])} '
                              f'ma bmi: {bmi(int(rekord[3]), float(rekord[4].strip('\n'))):.2f}')
            case '3':
                # rej, marka, model, nadwozie, paliwo, moc, pojemnosc, cena =\
                #     'SGLRF06', 'Chevrolet', 'Cruze', 'sedan', 'PB', 124, 1.6, 160000

                try:
                    root = pokaz_auta() # jeżeli są auta, to zostaną pokazane
                except:
                    root = Element('auta') # jeżeli nie mam aut to tworzę nowe

                rej = input('Podaj nr rej -> ')
                marka = input('Podaj marka -> ')
                model = input('Podaj model -> ')
                nadwozie = input('Podaj typ nadwozia -> ')
                paliwo = input('Podaj typ paliwa -> ')
                moc = input('Podaj moc w KM -> ')
                pojemnosc = input('Podaj poj silnika -> ')
                cena = input('Podaj cenę w PLN -> ')


                auto = SubElement(root, 'auto', rej=rej)
                SubElement(auto, 'marka').text = marka
                SubElement(auto, 'model').text = model
                SubElement(auto, 'nadwozie').text = nadwozie
                SubElement(auto, 'moc').text = str(moc)
                SubElement(auto, 'pojemnosc').text = f'{pojemnosc}'
                SubElement(auto, 'cena').text = '{}'.format(cena)

                # rej, marka, model, nadwozie, paliwo, moc, pojemnosc, cena = \
                #     'SG12345', 'Opel', 'Mokka', 'SUV', 'PB', 130, 1.2, 90000
                #
                # auto = SubElement(root, 'auto', rej=rej)
                # SubElement(auto, 'marka').text = marka
                # SubElement(auto, 'model').text = model
                # SubElement(auto, 'nadwozie').text = nadwozie
                # SubElement(auto, 'moc').text = str(moc)
                # SubElement(auto, 'pojemnosc').text = f'{pojemnosc}'
                # SubElement(auto, 'cena').text = '{}'.format(cena)

                # zapis
                ElementTree(root).write(r'files\auta.xml', encoding='utf-8')

            case '4':
                r = pokaz_auta()

                try:
                    print('Wyszukiwanie po rejestracji')
                    rej = input('Podaj nr rejestracji pojazdu -> ')
                    for a in r.findall('auto'):
                        if a.get('rej') == rej:
                            marka = a.find('marka').text
                            model = a.find('model').text
                            cena = a.find('cena').text
                            moc= a.find('moc').text
                            pojemnosc = a.find('pojemnosc').text
                            nadwozie = a.find('nadwozie').text
                            print(f'Pojazd marki: {marka}, {model}, kosztuje {cena} PLN, ma moc {moc} KM, nadwozie typu {nadwozie}')
                except: print('nie ma takiego pojazdu')

            case '5':
                r = pokaz_auta()

                try:
                    print('Wyszukiwanie po rejestracji')
                    rej = input('Podaj nr rejestracji pojazdu -> ')
                    for a in r.findall('auto'):
                        if a.get('rej') == rej:
                            odp = input('Czy zmieniamy rejestrację? (t-tak, n-nie) -> ')
                            if odp.upper() == 'T':
                                rej = input('podaj nowy numer rej -> ')
                                a.set('rej', rej)
                            odp = input('Czy korygujemy cenę? (t-tak, n-nie) -> ')
                            if odp.upper() == 'T':
                                cena = input('podaj nową cenę -> ')
                                a.find('cena').text = cena

                    ElementTree(r).write(r'files\auta.xml', encoding='utf-8')
                except:
                    print('nie ma takiego pojazdu')
            case '6':
                r = pokaz_auta()
                try:
                    print('Wyszukiwanie po rejestracji')
                    rej = input('Podaj nr rejestracji pojazdu -> ')
                    for a in r.findall('auto'):
                        if a.get('rej') == rej:
                            r.remove(a)

                    ElementTree(r).write(r'files\auta.xml', encoding='utf-8')
                except:
                    print('nie ma takiego pojazdu')
            case '7':
                import pickle
                print('marynowanie danych')
                przepis = ['marynowanie', 'kiszenie', 'kwaszenie']
                kształt = ['całe', 'talarki', 'różyczki', 'utarte', 'ćwiartki']
                producent = ['pudliszki', 'rolnik', 'babcia']

                plik = open(r'files\dane.bin', 'wb')
                pickle.dump(przepis, plik)
                pickle.dump(kształt, plik)
                pickle.dump(producent, plik)
                plik.close()

                print('otwieramy słoiki')
                p = open(r'files\dane.bin', 'rb')
                prz = pickle.load(p)
                ksz = pickle.load(p)
                pro = pickle.load(p)
                p.close()
                print('kształty: {0}'.format(ksz))
            case '8':
                import shelve
                print('pakowanie do szafki')
                plik = shelve.open(r'files\dane_sh.bin')
                plik['przepis'] = ['marynowanie', 'kiszenie', 'kwaszenie']
                plik['kształt'] = ['całe', 'talarki', 'różyczki', 'utarte', 'ćwiartki']
                plik['producent'] = ['pudliszki', 'rolnik', 'babcia']
                plik.sync() # zapis do pliku
                plik.close()

                print('otwieramy szafkę')
                plik_shelve = shelve.open(r'files\dane_sh.bin')
                print(plik_shelve['producent'])
                plik_shelve.close()

            case '9':
                from openpyxl import Workbook

                wb = Workbook() # utworzenie instancji workbook
                ws = wb.active # szukanie arkusza który teraz jest aktywny
                ws.title = 'dane osobowe'
                ws['A1'] = 'imie'
                ws['B1'] = 'nazwisko'
                ws['C1'] = 'pesel'
                ws['D1'] = 'waga'
                ws['E1'] = 'wzrost'

                # przykładowe dane
                ws['A2'] = 'Dariusz'
                ws['B2'] = 'Pieter'
                # pesel = input('Podaj pesel -> ')
                pesel = '09282385753'
                if validate_pesel(pesel): ws['C2'] = "'" + pesel
                else: raise ArgumentError('to nie jest poprawny pesel')
                ws['D2'] = 102
                ws['E2'] = 1.87

                wb.save(r'files\dane_osobowe.xlsx')
                wb.close()

                # dodanie kolejnej osoby do excela
                from openpyxl import load_workbook
                wb=load_workbook(r"files\dane_osobowe.xlsx")
                ws = wb['dane osobowe'] # wskazanie na arkusz z którym chemy pracować
                nr_wiersza = ws.max_row # szukam ostatniego zajętego wiersza
                nr_wiersza += 1

                ws.cell(nr_wiersza, column=1, value='Kacper')
                ws.cell(nr_wiersza, column=2, value='Karpała')
                ws.cell(nr_wiersza, column=3, value="'" + '01322389949')
                ws.cell(nr_wiersza, column=4, value=78)
                ws.cell(nr_wiersza, column=5, value=1.80)

                wb.save(r'files\dane_osobowe.xlsx')
                wb.close()

            case _: break





















